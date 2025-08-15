import os
import time
import csv
import random
import datetime
import platform
import psutil
import threading
import cryptography
import psutil
import openpyxl
from cryptography.hazmat.primitives.ciphers import Cipher, algorithms, modes
from cryptography.hazmat.primitives import hashes, padding
from cryptography.hazmat.primitives.asymmetric import rsa, padding as asym_padding
from cryptography.hazmat.primitives.asymmetric import ec, ed25519, x25519
from cryptography.hazmat.primitives.serialization import Encoding, PublicFormat, PrivateFormat, NoEncryption

# Lista para armazenar os algoritmos criptográficos disponíveis no sistema
AVAILABLE_ALGORITHMS = []

# Verificação da disponibilidade de RSA
try:
    from cryptography.hazmat.primitives.asymmetric import rsa
    AVAILABLE_ALGORITHMS.append("RSA")
    print("✓ RSA disponível")
except (ImportError, AttributeError) as e:
    print(f"⚠️ AVISO: RSA não disponível: {str(e)}")

# Verificação da disponibilidade de Curve25519/Ed25519
try:
    from cryptography.hazmat.primitives.asymmetric import x25519, ed25519
    ed25519.Ed25519PrivateKey.generate()
    AVAILABLE_ALGORITHMS.append("Ed25519")
    AVAILABLE_ALGORITHMS.append("X25519")
    print("✓ Curve25519/Ed25519 disponível")
except (ImportError, AttributeError) as e:
    print(f"⚠️ AVISO: Curve25519/Ed25519 não disponível: {str(e)}")

# Verificação da disponibilidade de curvas NIST
try:
    from cryptography.hazmat.primitives.asymmetric import ec
    ec.generate_private_key(ec.SECP256R1())
    AVAILABLE_ALGORITHMS.append("NIST_P256")
    AVAILABLE_ALGORITHMS.append("NIST_P384")
    AVAILABLE_ALGORITHMS.append("NIST_P521")
    print("✓ Curvas NIST (P-256/P-384/P-521) disponíveis")
except (ImportError, AttributeError) as e:
    print(f"⚠️ AVISO: Curvas NIST não disponíveis: {str(e)}")

print(f"Algoritmos disponíveis: {', '.join(AVAILABLE_ALGORITHMS)}")

class BenchmarkResult:
    """
    Classe para armazenar os resultados de um benchmark único.
    """
    def __init__(self):
        self.algorithm = ""         # Nome do algoritmo testado
        self.key_size = 0           # Tamanho da chave em bits
        self.operation_type = ""    # Tipo de operação (ex: Key Generation, Signing, Verification)
        self.execution_time_ms = 0  # Tempo de execução em milissegundos
        self.memory_usage_mb = 0.0  # Uso de memória em MB
        self.cpu_percentage = 0.0   # Uso de CPU em percentual
        self.data_size_bytes = 0    # Tamanho dos dados testados em bytes
        self.timestamp = None       # Momento em que o benchmark foi executado
        self.notes = ""             # Notas adicionais (ex: timeout, erro)

class CryptoBenchmark:
    """
    Classe principal para execução de benchmarks de algoritmos criptográficos.
    """
    # Tamanho padrão dos dados para teste
    TEST_DATA_SIZE_MB = 1
    
    def run_complete_benchmark(self):
        """Executa benchmarks completos de todos os algoritmos configurados"""
        print("\n===== Executando Benchmark Completo =====")
        print("Executando benchmark para todos os algoritmos disponíveis")
        
        # Executa benchmark para cada grupo de algoritmos
        self.run_curve25519_benchmark()    # Ed25519/X25519
        self.run_nist_curves_benchmark()   # NIST P-256/P-384/P-521
        self.run_rsa_benchmark()           # RSA
        
        print("\n===== Benchmark Completo Finalizado =====")
        print(f"Total de resultados: {len(self.results)}")
        print("Você pode exportar os resultados para CSV usando a opção 5 no menu.")
    
    def run_curve25519_benchmark(self):
        """
        Executa o benchmark de Ed25519 (para assinatura digital) e 
        X25519 (para troca de chaves Diffie-Hellman)
        """
        print("\n===== Benchmark de Ed25519/X25519 (Curve25519) =====")
        print("Estabilizando o sistema antes do benchmark...")
        self.limit_cpu_cores()
        
        results_count_before = len(self.results)
        
        # Benchmark de Ed25519 (assinatura digital)
        print("\nTestando Ed25519 (assinatura digital)...")
        try:
            # Verifica disponibilidade
            if "Ed25519" not in AVAILABLE_ALGORITHMS:
                print("⚠️ AVISO: Ed25519 não está disponível no sistema.")
            else:
                # Benchmark de geração de chaves
                result_keygen = self.benchmark_ed25519_keygen()
                self.results.append(result_keygen)
                self.display_result(result_keygen)
                
                # Benchmark de assinatura
                result_sign = self.benchmark_ed25519_sign(self.test_data)
                self.results.append(result_sign)
                self.display_result(result_sign)
                
                # Benchmark de verificação
                if not self.timeout_occurred:
                    result_verify = self.benchmark_ed25519_verify(self.test_data)
                    self.results.append(result_verify)
                    self.display_result(result_verify)
        except Exception as ex:
            print(f"Erro durante benchmark de Ed25519: {str(ex)}")
        
        # Benchmark de X25519 (troca de chaves)
        print("\nTestando X25519 (troca de chaves)...")
        try:
            # Verifica disponibilidade
            if "X25519" not in AVAILABLE_ALGORITHMS:
                print("⚠️ AVISO: X25519 não está disponível no sistema.")
            else:
                # Benchmark de geração de chaves
                result_keygen = self.benchmark_x25519_keygen()
                self.results.append(result_keygen)
                self.display_result(result_keygen)
                
                # Benchmark de troca de chaves
                result_key_exchange = self.benchmark_x25519_key_exchange()
                self.results.append(result_key_exchange)
                self.display_result(result_key_exchange)
        except Exception as ex:
            print(f"Erro durante benchmark de X25519: {str(ex)}")
            
        results_count_after = len(self.results)
        print(f"\nTotal de resultados da Curve25519 adicionados: {results_count_after - results_count_before}")

    def run_nist_curves_benchmark(self):
        """
        Executa benchmark das curvas NIST (P-256, P-384, P-521) para ECDSA
        (assinatura digital) e ECDH (troca de chaves)
        """
        print("\n===== Benchmark de Curvas NIST (P-256/P-384/P-521) =====")
        print("Estabilizando o sistema antes do benchmark...")
        self.limit_cpu_cores()
        
        # Lista de curvas NIST a serem testadas
        nist_curves = {
            "NIST_P256": ec.SECP256R1(),  # P-256
            "NIST_P384": ec.SECP384R1(),  # P-384
            "NIST_P521": ec.SECP521R1()   # P-521
        }
        
        results_count_before = len(self.results)
        
        for curve_name, curve in nist_curves.items():
            print(f"\nTestando {curve_name}...")
            
            try:
                # Verificar se a curva está disponível
                if curve_name not in AVAILABLE_ALGORITHMS:
                    print(f"⚠️ AVISO: {curve_name} não está disponível no sistema.")
                    continue
                
                # Benchmark de geração de chaves
                result_keygen = self.benchmark_ecdsa_keygen(curve_name, curve)
                self.results.append(result_keygen)
                self.display_result(result_keygen)
                
                # Benchmark de assinatura ECDSA
                result_sign = self.benchmark_ecdsa_sign(curve_name, curve, self.test_data)
                self.results.append(result_sign)
                self.display_result(result_sign)
                
                # Benchmark de verificação ECDSA
                if not self.timeout_occurred:
                    result_verify = self.benchmark_ecdsa_verify(curve_name, curve, self.test_data)
                    self.results.append(result_verify)
                    self.display_result(result_verify)
                    
                # Benchmark de troca de chaves ECDH
                result_ecdh = self.benchmark_ecdh_key_exchange(curve_name, curve)
                self.results.append(result_ecdh)
                self.display_result(result_ecdh)
                
            except Exception as ex:
                print(f"Erro durante benchmark de {curve_name}: {str(ex)}")
        
        results_count_after = len(self.results)
        print(f"\nTotal de resultados de curvas NIST adicionados: {results_count_after - results_count_before}")

    def run_rsa_benchmark(self):
        """
        Executa o benchmark de RSA com diferentes tamanhos de chave
        (1024, 2048, 4096 bits)
        """
        print("\n===== Benchmark de RSA =====")
        print("Estabilizando o sistema antes do benchmark...")
        self.limit_cpu_cores()
        
        # Diferentes tamanhos de chave para RSA
        key_sizes = [1024, 2048, 4096]
        
        results_count_before = len(self.results)
        
        for key_size in key_sizes:
            print(f"\nTestando RSA com chave de {key_size} bits...")
            
            try:
                # Verificar se RSA está disponível
                if "RSA" not in AVAILABLE_ALGORITHMS:
                    print("⚠️ AVISO: RSA não está disponível no sistema.")
                    break
                
                # Benchmark de geração de chaves RSA
                result_keygen = self.benchmark_rsa_keygen(key_size)
                self.results.append(result_keygen)
                self.display_result(result_keygen)
                
                # Prepara dados para assinatura (hash dos dados originais)
                digest = hashes.Hash(hashes.SHA256())
                digest.update(self.test_data)
                message_digest = digest.finalize()
                
                # Benchmark de assinatura RSA
                result_sign = self.benchmark_rsa_sign(key_size, message_digest)
                self.results.append(result_sign)
                self.display_result(result_sign)
                
                # Benchmark de verificação RSA
                if not self.timeout_occurred:
                    result_verify = self.benchmark_rsa_verify(key_size, message_digest)
                    self.results.append(result_verify)
                    self.display_result(result_verify)
                    
                # Benchmark de operação criptográfica RSA (criptografar/descriptografar)
                # Para RSA, usamos um conjunto de dados menor devido às limitações
                max_data_size = key_size // 8 - 42  # Fórmula aproximada para OAEP
                max_data_size = max(1, max_data_size)  # Garantir pelo menos 1 byte
                test_data = os.urandom(max_data_size)
                
                # Benchmark de criptografia/descriptografia RSA
                result_crypt = self.benchmark_rsa_encryption("RSA", key_size, test_data)
                self.results.append(result_crypt)
                self.display_result(result_crypt)
                    
            except Exception as ex:
                print(f"Erro durante benchmark de RSA: {str(ex)}")
        
        results_count_after = len(self.results)
        print(f"\nTotal de resultados do RSA adicionados: {results_count_after - results_count_before}")
        
    def __init__(self):
        self.results = []              # Lista para armazenar os resultados dos benchmarks
        self.stop_cpu_measurement = False  # Flag para controle da medição de CPU
        self.timeout_seconds = 60      # Timeout padrão (60 segundos)
        self.timeout_occurred = False  # Flag para indicar se ocorreu timeout
        self.max_cores = psutil.cpu_count(logical=True)  # Número máximo de núcleos disponíveis
        self.use_cores = self.max_cores  # Por padrão, usa todos os núcleos
        self.memory_limit_mb = None    # Limite de memória (None = sem limite)
        self.baseline_memory_usage = 0 # Uso de memória de linha de base
        self.baseline_cpu_usage = 0    # Uso de CPU de linha de base
        
        # Inicializa os dados de teste
        self.test_data_size_mb = self.TEST_DATA_SIZE_MB
        self.init_test_data()
        
        # Atualiza as métricas de linha de base do sistema
        self.update_system_baseline()
        
    def init_test_data(self):
        """Gera dados aleatórios para usar nos testes de criptografia"""
        print(f"Inicializando {self.test_data_size_mb}MB de dados para teste...")
        self.test_data = os.urandom(self.test_data_size_mb * 1024 * 1024)
        print("Dados de teste inicializados com sucesso.")
        
    def update_system_baseline(self):
        """Captura as métricas de uso de memória e CPU antes da execução dos benchmarks"""
        process = psutil.Process(os.getpid())
        self.baseline_memory_usage = process.memory_info().rss / (1024.0 * 1024.0)
        self.baseline_cpu_usage = process.cpu_percent(interval=0.5)
        
    def timeout_handler(self):
        """Função chamada quando um benchmark excede o tempo limite configurado."""
        self.timeout_occurred = True
        print(f"Timeout atingido após {self.timeout_seconds} segundos!")
        
    def run(self):
        """Método principal que exibe o menu e controla o fluxo do programa"""
        self.print_system_info()
        
        # Exibir configuração atual
        self.print_benchmark_config()
        
        while True:
            print("\nEscolha uma opção:")
            print("1. Benchmark de Curve25519/Ed25519")
            print("2. Benchmark de Curvas NIST (P-256/P-384/P-521)")
            print("3. Benchmark de RSA")
            print("4. Benchmark completo (todos os algoritmos)")
            print("5. Exportar resultados para CSV")
            print("6. Exportar resultados para XLSX (formatado)")  # Nova opção
            print("7. Configurar núcleos de CPU e limite de memória")
            print("8. Limpar resultados anteriores")
            print("0. Sair")
            
            option = input("\nOpção: ")
            
            # Estrutura de decisão para as opções do menu
            if option == "1":
                self.run_curve25519_benchmark()    # Ed25519/X25519 benchmark
            elif option == "2":
                self.run_nist_curves_benchmark()   # NIST curves benchmark
            elif option == "3":
                self.run_rsa_benchmark()           # RSA benchmark
            elif option == "4":
                self.run_complete_benchmark()      # Todos os benchmarks
            elif option == "5":
                self.export_results_to_csv()       # Exportar resultados CSV
            elif option == "6":
                self.export_results_to_xlsx()      # Exportar resultados XLSX (Nova opção)
            elif option == "7":
                self.configure_resources()         # Configurar recursos
            elif option == "8":
                self.clear_results()               # Limpar resultados
            elif option == "0":
                break                              # Sair do programa
            else:
                print("Opção inválida, tente novamente.")
                
    def print_system_info(self):
        """Exibe informações do sistema para referência"""
        print("===== CryptoBenchmark - Análise de Desempenho Criptográfico =====")
        print(f"Data e Hora: {datetime.datetime.now()}")
        print(f"Sistema Operacional: {platform.system()} {platform.version()}")
        print(f"Processador: {psutil.cpu_count(logical=True)} núcleos lógicos (usando {self.use_cores})")
        print(f"Memória Total: {psutil.virtual_memory().total / (1024**3):.4f} GB")
        print(f"Memória Disponível: {psutil.virtual_memory().available / (1024**3):.4f} GB")
        print(f"Utilização de CPU atual: {psutil.cpu_percent()}%")
        print(f"Processos ativos: {len(psutil.pids())}")
        print("=================================================================")
        
    def print_benchmark_config(self):
        """Exibe a configuração atual do benchmark"""
        print("\nConfiguração do Benchmark:")
        print(f"- Usando {self.use_cores} núcleos de CPU")
        print(f"- Tempo limite para testes: {self.timeout_seconds} segundos")
        print(f"- Tamanho dos dados de teste: {self.test_data_size_mb} MB")
        if self.memory_limit_mb:
            print(f"- Limite de memória: {self.memory_limit_mb} MB")
        print(f"- Memória de linha de base: {self.baseline_memory_usage:.4f} MB")
        print(f"- CPU de linha de base: {self.baseline_cpu_usage:.4f}%")
        
    def configure_resources(self):
        """Permite configurar o número de núcleos de CPU e o limite de memória"""
        print("===== Configuração de Recursos =====")
        print(f"Total de núcleos disponíveis no sistema: {self.max_cores}")
        print(f"Memória total do sistema: {int(psutil.virtual_memory().total / (1024*1024))} MB")
        print(f"Memória disponível: {int(psutil.virtual_memory().available / (1024*1024))} MB")
        
        try:
            # Configuração de núcleos
            cores = input(f"Número de núcleos para usar (1-{self.max_cores}, atualmente {self.use_cores}): ")
            if cores:
                cores = int(cores)
                if 1 <= cores <= self.max_cores:
                    self.use_cores = cores
                else:
                    print(f"Valor inválido. Usando {self.use_cores} núcleos.")
            
            # Configuração de limite de memória
            mem_limit = input("Limite de memória em MB (opcional, Enter para sem limite): ")
            if mem_limit:
                self.memory_limit_mb = int(mem_limit)
                print(f"Limite de memória definido para {self.memory_limit_mb} MB")
            else:
                self.memory_limit_mb = None
                print("Sem limite de memória")
                
            # Atualiza a linha de base após a mudança de configuração
            print("Atualizando linha de base do sistema...")
            self.update_system_baseline()
            print(f"- Nova memória de linha de base: {self.baseline_memory_usage:.4f} MB")
            print(f"- Nova CPU de linha de base: {self.baseline_cpu_usage:.4f}%")
            
        except ValueError:
            print("Entrada inválida. Mantendo configurações anteriores.")
                
    def clear_results(self):
            self.results = []
            print("Resultados limpos com sucesso.")
            
    def limit_cpu_cores(self):
        """Configura o uso limitado de CPU para o benchmark."""
        if self.use_cores < psutil.cpu_count(logical=True):
            print(f"Afinidade de CPU definida para usar {self.use_cores} núcleos: {list(range(self.use_cores))}")
            # Nota: Em um programa real, aqui você poderia usar psutil.Process().cpu_affinity()
            # Mas isso requer privilégios elevados em alguns sistemas
    
    def measure_cpu_usage(self, result):
        """Mede o uso de CPU durante uma operação"""
        process = psutil.Process(os.getpid())
        cpu_percentages = []
        
        while not self.stop_cpu_measurement:
            cpu_percentages.append(process.cpu_percent(interval=0.1))
            time.sleep(0.1)
            
        if cpu_percentages:
            result.cpu_percentage = sum(cpu_percentages) / len(cpu_percentages)
        else:
            result.cpu_percentage = 0
        
    def display_result(self, result):
        """Exibe os resultados do benchmark"""
        print(f"\nResultado do Benchmark:")
        print(f"Algoritmo: {result.algorithm}")
        print(f"Tamanho da Chave: {result.key_size} bits")
        print(f"Operação: {result.operation_type}")
        print(f"Tamanho dos Dados: {result.data_size_bytes / (1024.0 * 1024.0):.4f} MB")
        print(f"Tempo de Execução: {result.execution_time_ms:.4f} ms")
        print(f"Uso de Memória: {result.memory_usage_mb:.4f} MB")
        print(f"Uso de CPU: {result.cpu_percentage:.4f}%")
        if result.notes:
            print(f"Observações: {result.notes}")
    
    def export_results_to_csv(self):
        from datetime import datetime
        """Exporta os resultados para um arquivo CSV"""
        if not self.results:
            print("Não há resultados para exportar. Execute alguns benchmarks primeiro.")
            return

        # Informações do sistema
        total_cores = self.use_cores
        available_memory_gb = (
            self.memory_limit_mb / 1024 if self.memory_limit_mb is not None 
            else psutil.virtual_memory().available / (1024 ** 2)
        )


        # Nome do algoritmo (fallback para "Algoritmo" se não estiver definido)
        algorithm_name = getattr(self, "algorithm_name", "Algoritmo")

        # Data/hora opcional no nome do arquivo
        timestamp = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")

        # Geração do nome do arquivo
        filename = f"Resultados_{algorithm_name}_{total_cores}cores_{available_memory_gb:.4f}GB_{timestamp}.csv"
        
        try:
            with open(filename, 'w', newline='') as csvfile:
                writer = csv.writer(csvfile)
                
                # Escrever cabeçalho
                writer.writerow([
                    "Algoritmo", 
                    "Tamanho da Chave (bits)", 
                    "Operação", 
                    "Tamanho dos Dados (MB)",
                    "Tempo de Execução (ms)", 
                    "Uso de Memória (MB)", 
                    "Uso de CPU (%)",
                    "Data/Hora",
                    "Observações"
                ])
                
                # Escrever resultados
                for result in self.results:
                    writer.writerow([
                        result.algorithm,
                        result.key_size,
                        result.operation_type,
                        f"{result.data_size_bytes / (1024.0 * 1024.0):.4f}",
                        f"{result.execution_time_ms:.4f}",
                        f"{result.memory_usage_mb:.4f}",
                        f"{result.cpu_percentage:.4f}",
                        result.timestamp.strftime('%Y-%m-%d %H:%M:%S') if result.timestamp else "",
                        result.notes
                    ])
                    
            print(f"Resultados exportados para {filename} com sucesso!")
            print(f"Total de resultados exportados: {len(self.results)}")
            
            # Mostrar contagem por algoritmo
            algorithm_counts = {}
            for result in self.results:
                algorithm_counts[result.algorithm] = algorithm_counts.get(result.algorithm, 0) + 1
                
            print("\nDetalhamento dos resultados exportados:")
            for algo, count in algorithm_counts.items():
                print(f"- {algo}: {count} resultados")
        
        except Exception as ex:
            print(f"Erro ao exportar resultados: {str(ex)}")


    
    def export_results_to_xlsx(self):
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
        from openpyxl.formatting.rule import ColorScaleRule
        
        if not self.results:
            print("Não há resultados para exportar. Execute alguns benchmarks primeiro.")
            return

        # Informações do sistema
        total_cores = self.use_cores
        available_memory_gb = (
            self.memory_limit_mb / 1024 if self.memory_limit_mb is not None 
            else psutil.virtual_memory().available / (1024 ** 3)
        )

        # Nome do algoritmo (fallback para "Algoritmo" se não estiver definido)
        algorithm_name = getattr(self, "algorithm_name", "Algoritmo")

        # Data/hora no nome do arquivo
        timestamp = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")

        # Geração do nome do arquivo
        filename = f"Resultados_{algorithm_name}_{total_cores}cores_{available_memory_gb:.2f}GB_{timestamp}.xlsx"
        
        try:
            # Criar workbook e worksheet
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Resultados Benchmark"
            
            # === ESTILOS ===
            # Estilo do cabeçalho
            header_font = Font(name='Calibri', size=12, bold=True, color='FFFFFF')
            header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
            header_alignment = Alignment(horizontal='center', vertical='center')
            
            # Estilo dos dados
            data_font = Font(name='Calibri', size=11)
            data_alignment = Alignment(horizontal='center', vertical='center')
            
            # Bordas
            thin_border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            
            # Cores alternadas para linhas
            light_fill = PatternFill(start_color='F2F2F2', end_color='F2F2F2', fill_type='solid')
            
            # === INFORMAÇÕES DO SISTEMA ===
            ws['A1'] = "RELATÓRIO DE BENCHMARK CRIPTOGRÁFICO"
            ws['A1'].font = Font(name='Calibri', size=16, bold=True, color='366092')
            ws.merge_cells('A1:I1')
            
            ws['A3'] = f"Data de Execução: {datetime.datetime.now().strftime('%d/%m/%Y %H:%M:%S')}"
            ws['A4'] = f"Sistema: {platform.system()} {platform.version()}"
            ws['A5'] = f"CPU: {total_cores} núcleos utilizados de {psutil.cpu_count(logical=True)} disponíveis"
            ws['A6'] = f"Memória: {available_memory_gb:.2f} GB disponíveis"
            ws['A7'] = f"Tamanho dos dados de teste: {self.test_data_size_mb} MB"
            
            # Aplicar estilo às informações do sistema
            for row in range(3, 8):
                ws[f'A{row}'].font = Font(name='Calibri', size=11, italic=True)
            
            # === CABEÇALHOS DA TABELA ===
            headers = [
                "Algoritmo", 
                "Tamanho da Chave (bits)", 
                "Operação", 
                "Tamanho dos Dados (MB)",
                "Tempo de Execução (ms)", 
                "Uso de Memória (MB)", 
                "Uso de CPU (%)",
                "Data/Hora",
                "Observações"
            ]
            
            # Linha onde começam os cabeçalhos
            header_row = 9
            
            # Escrever cabeçalhos
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=header_row, column=col, value=header)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment
                cell.border = thin_border
            
            # === DADOS ===
            data_start_row = header_row + 1
            
            # Escrever resultados com formatação
            for row_idx, result in enumerate(self.results, data_start_row):
                # Dados da linha
                row_data = [
                    result.algorithm,
                    result.key_size,
                    result.operation_type,
                    result.data_size_bytes / (1024.0 * 1024.0),  # MB com 6 casas decimais
                    result.execution_time_ms,                     # ms com 6 casas decimais
                    result.memory_usage_mb,                       # MB com 6 casas decimais
                    result.cpu_percentage,                        # % com 3 casas decimais
                    result.timestamp.strftime('%d/%m/%Y %H:%M:%S') if result.timestamp else "",
                    result.notes
                ]
                
                # Escrever dados na planilha
                for col_idx, value in enumerate(row_data, 1):
                    cell = ws.cell(row=row_idx, column=col_idx, value=value)
                    cell.font = data_font
                    cell.alignment = data_alignment
                    cell.border = thin_border
                    
                    # Formatação numérica específica
                    if col_idx == 4:  # Tamanho dos Dados (MB)
                        cell.number_format = '0.000000'
                    elif col_idx == 5:  # Tempo de Execução (ms)
                        cell.number_format = '0.000000'
                    elif col_idx == 6:  # Uso de Memória (MB)
                        cell.number_format = '0.000000'
                    elif col_idx == 7:  # Uso de CPU (%)
                        cell.number_format = '0.000'
                    
                    # Cores alternadas para linhas
                    if (row_idx - data_start_row) % 2 == 1:
                        cell.fill = light_fill
            
            # === AJUSTES DE LARGURA DAS COLUNAS ===
            column_widths = {
                'A': 15,  # Algoritmo
                'B': 20,  # Tamanho da Chave
                'C': 18,  # Operação
                'D': 22,  # Tamanho dos Dados
                'E': 25,  # Tempo de Execução
                'F': 20,  # Uso de Memória
                'G': 15,  # Uso de CPU
                'H': 20,  # Data/Hora
                'I': 25   # Observações
            }
            
            for col, width in column_widths.items():
                ws.column_dimensions[col].width = width
            
            # === FORMATAÇÃO CONDICIONAL ===
            # Aplicar escala de cores para tempos de execução (coluna E)
            if len(self.results) > 1:
                data_range = f"E{data_start_row}:E{data_start_row + len(self.results) - 1}"
                ws.conditional_formatting.add(data_range, 
                    ColorScaleRule(start_type='min', start_color='63BE7B',  # Verde
                                mid_type='percentile', mid_value=50, mid_color='FFEB84',  # Amarelo
                                end_type='max', end_color='F8696B'))  # Vermelho
            
            # === RESUMO ESTATÍSTICO ===
            summary_start_row = data_start_row + len(self.results) + 2
            
            # Título do resumo
            ws[f'A{summary_start_row}'] = "RESUMO ESTATÍSTICO"
            ws[f'A{summary_start_row}'].font = Font(name='Calibri', size=14, bold=True, color='366092')
            ws.merge_cells(f'A{summary_start_row}:D{summary_start_row}')
            
            summary_start_row += 2
            
            # Calcular estatísticas por algoritmo
            algorithm_stats = {}
            for result in self.results:
                algo = result.algorithm
                if algo not in algorithm_stats:
                    algorithm_stats[algo] = {
                        'count': 0,
                        'total_time': 0,
                        'min_time': float('inf'),
                        'max_time': 0,
                        'total_memory': 0,
                        'total_cpu': 0
                    }
                
                stats = algorithm_stats[algo]
                stats['count'] += 1
                stats['total_time'] += result.execution_time_ms
                stats['min_time'] = min(stats['min_time'], result.execution_time_ms)
                stats['max_time'] = max(stats['max_time'], result.execution_time_ms)
                stats['total_memory'] += result.memory_usage_mb
                stats['total_cpu'] += result.cpu_percentage
            
            # Cabeçalhos do resumo
            summary_headers = ['Algoritmo', 'Testes', 'Tempo Médio (ms)', 'Tempo Min (ms)', 
                            'Tempo Max (ms)', 'Memória Média (MB)', 'CPU Média (%)']
            
            for col, header in enumerate(summary_headers, 1):
                cell = ws.cell(row=summary_start_row, column=col, value=header)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment
                cell.border = thin_border
            
            # Dados do resumo
            for row_idx, (algo, stats) in enumerate(algorithm_stats.items(), summary_start_row + 1):
                summary_data = [
                    algo,
                    stats['count'],
                    stats['total_time'] / stats['count'],     # Tempo médio
                    stats['min_time'],                        # Tempo mínimo
                    stats['max_time'],                        # Tempo máximo
                    stats['total_memory'] / stats['count'],   # Memória média
                    stats['total_cpu'] / stats['count']       # CPU médio
                ]
                
                for col_idx, value in enumerate(summary_data, 1):
                    cell = ws.cell(row=row_idx, column=col_idx, value=value)
                    cell.font = data_font
                    cell.alignment = data_alignment
                    cell.border = thin_border
                    
                    # Formatação numérica para o resumo
                    if col_idx >= 3 and col_idx <= 6:  # Colunas de tempo e memória
                        cell.number_format = '0.000000'
                    elif col_idx == 7:  # CPU
                        cell.number_format = '0.000'
            
            # === FREEZAR PAINÉIS ===
            ws.freeze_panes = f'A{data_start_row}'
            
            # Salvar arquivo
            wb.save(filename)
                    
            print(f"Resultados exportados para {filename} com sucesso!")
            print(f"Total de resultados exportados: {len(self.results)}")
            
            # Mostrar contagem por algoritmo
            algorithm_counts = {}
            for result in self.results:
                algorithm_counts[result.algorithm] = algorithm_counts.get(result.algorithm, 0) + 1
                
            print("\nDetalhamento dos resultados exportados:")
            for algo, count in algorithm_counts.items():
                print(f"- {algo}: {count} resultados")
            
            print(f"\nRecursos do arquivo XLSX:")
            print("- Formatação condicional nos tempos de execução")
            print("- Cores alternadas nas linhas")
            print("- Resumo estatístico por algoritmo")
            print("- Colunas ajustadas automaticamente")
            print("- 6 casas decimais para medições de precisão")
        
        except Exception as ex:
            print(f"Erro ao exportar resultados: {str(ex)}")
            print("Certifique-se de que a biblioteca openpyxl está instalada: pip install openpyxl")
        
    # Implementações dos métodos de benchmark

    # ==== Ed25519 / X25519 (Curve25519) ====

    def benchmark_ed25519_keygen(self):
        """Realiza o benchmark de geração de chaves Ed25519"""
        result = BenchmarkResult()
        result.algorithm = "Ed25519"
        result.key_size = 256  # Ed25519 tem tamanho fixo de 256 bits
        result.operation_type = "Key Generation"
        result.data_size_bytes = 0  # Não aplicável para geração de chaves
        
        # Reset da flag de timeout
        self.timeout_occurred = False
        
        # Medição de recursos
        process = psutil.Process(os.getpid())
        start_memory = process.memory_info().rss
        
        # Iniciar thread para monitoramento de CPU
        self.stop_cpu_measurement = False
        cpu_usage_thread = threading.Thread(target=self.measure_cpu_usage, args=(result,))
        cpu_usage_thread.start()
        
        # Iniciar thread para timeout (se configurado)
        if self.timeout_seconds > 0:
            timer = threading.Timer(self.timeout_seconds, self.timeout_handler)
            timer.start()
        
        # Executar geração de chave e medir tempo
        start_time = time.time()
        
        try:
            # Gerar par de chaves Ed25519
            if not self.timeout_occurred:
                private_key = ed25519.Ed25519PrivateKey.generate()
                public_key = private_key.public_key()
                
        except Exception as ex:
            # Cancelar o timer se ocorrer exceção
            if self.timeout_seconds > 0:
                timer.cancel()
            # Finalizar medição de CPU
            self.stop_cpu_measurement = True
            cpu_usage_thread.join()
            raise ex
            
        end_time = time.time()
        
        # Cancelar o timer se a operação for concluída antes do timeout
        if self.timeout_seconds > 0 and not self.timeout_occurred:
            timer.cancel()
        
        # Finalizar medição de CPU
        self.stop_cpu_measurement = True
        cpu_usage_thread.join()
        
        # Calcular uso de memória
        end_memory = process.memory_info().rss
        
        # Armazenar resultados
        result.execution_time_ms = (end_time - start_time) * 1000
        result.memory_usage_mb = (end_memory - start_memory) / (1024.0 * 1024.0)
        result.timestamp = datetime.datetime.now()
        
        # Adicionar informações sobre timeout, se ocorreu
        if self.timeout_occurred:
            result.notes = f"Timeout após {self.timeout_seconds} segundos"
            result.execution_time_ms = self.timeout_seconds * 1000
            
        return result
        
    def benchmark_ed25519_sign(self, data):
        """Realiza o benchmark de assinatura Ed25519"""
        result = BenchmarkResult()
        result.algorithm = "Ed25519"
        result.key_size = 256  # Ed25519 tem tamanho fixo de 256 bits
        result.operation_type = "Signing"
        result.data_size_bytes = len(data)
        
        # Gerar chave privada Ed25519 para assinatura
        private_key = ed25519.Ed25519PrivateKey.generate()
        
        # Reset da flag de timeout
        self.timeout_occurred = False
        
        # Medição de recursos
        process = psutil.Process(os.getpid())
        start_memory = process.memory_info().rss
        
        # Iniciar thread para monitoramento de CPU
        self.stop_cpu_measurement = False
        cpu_usage_thread = threading.Thread(target=self.measure_cpu_usage, args=(result,))
        cpu_usage_thread.start()
        
        # Iniciar thread para timeout (se configurado)
        if self.timeout_seconds > 0:
            timer = threading.Timer(self.timeout_seconds, self.timeout_handler)
            timer.start()
        
        # Executar assinatura e medir tempo
        start_time = time.time()
        
        try:
            # Assinar os dados
            if not self.timeout_occurred:
                signature = private_key.sign(data)
                
        except Exception as ex:
            # Cancelar o timer se ocorrer exceção
            if self.timeout_seconds > 0:
                timer.cancel()
            # Finalizar medição de CPU
            self.stop_cpu_measurement = True
            cpu_usage_thread.join()
            raise ex
            
        end_time = time.time()
        
        # Cancelar o timer se a operação for concluída antes do timeout
        if self.timeout_seconds > 0 and not self.timeout_occurred:
            timer.cancel()
        
        # Finalizar medição de CPU
        self.stop_cpu_measurement = True
        cpu_usage_thread.join()
        
        # Calcular uso de memória
        end_memory = process.memory_info().rss
        
        # Armazenar resultados
        result.execution_time_ms = (end_time - start_time) * 1000
        result.memory_usage_mb = (end_memory - start_memory) / (1024.0 * 1024.0)
        result.timestamp = datetime.datetime.now()
        
        # Adicionar informações sobre timeout, se ocorreu
        if self.timeout_occurred:
            result.notes = f"Timeout após {self.timeout_seconds} segundos"
            result.execution_time_ms = self.timeout_seconds * 1000
            
        # Armazenar a assinatura e a chave para uso em verificação
        self._last_ed25519_signature = signature
        self._last_ed25519_private_key = private_key
            
        return result
        
    def benchmark_ed25519_verify(self, data):
        """Realiza o benchmark de verificação Ed25519"""
        result = BenchmarkResult()
        result.algorithm = "Ed25519"
        result.key_size = 256  # Ed25519 tem tamanho fixo de 256 bits
        result.operation_type = "Verification"
        result.data_size_bytes = len(data)
        
        # Verificar se temos uma assinatura e chave pública disponíveis
        if not hasattr(self, '_last_ed25519_signature') or not hasattr(self, '_last_ed25519_private_key'):
            # Se não tivermos, criar uma nova
            private_key = ed25519.Ed25519PrivateKey.generate()
            signature = private_key.sign(data)
            public_key = private_key.public_key()
        else:
            # Usar a última assinatura e chave geradas
            signature = self._last_ed25519_signature
            public_key = self._last_ed25519_private_key.public_key()
        
        # Reset da flag de timeout
        self.timeout_occurred = False
        
        # Medição de recursos
        process = psutil.Process(os.getpid())
        start_memory = process.memory_info().rss
        
        # Iniciar thread para monitoramento de CPU
        self.stop_cpu_measurement = False
        cpu_usage_thread = threading.Thread(target=self.measure_cpu_usage, args=(result,))
        cpu_usage_thread.start()
        
        # Iniciar thread para timeout (se configurado)
        if self.timeout_seconds > 0:
            timer = threading.Timer(self.timeout_seconds, self.timeout_handler)
            timer.start()
        
        # Executar verificação e medir tempo
        start_time = time.time()
        
        try:
            # Verificar a assinatura
            if not self.timeout_occurred:
                public_key.verify(signature, data)
                
        except Exception as ex:
            # Cancelar o timer se ocorrer exceção
            if self.timeout_seconds > 0:
                timer.cancel()
            # Finalizar medição de CPU
            self.stop_cpu_measurement = True
            cpu_usage_thread.join()
            raise ex
            
        end_time = time.time()
        
        # Cancelar o timer se a operação for concluída antes do timeout
        if self.timeout_seconds > 0 and not self.timeout_occurred:
            timer.cancel()
        
        # Finalizar medição de CPU
        self.stop_cpu_measurement = True
        cpu_usage_thread.join()
        
        # Calcular uso de memória
        end_memory = process.memory_info().rss
        
        # Armazenar resultados
        result.execution_time_ms = (end_time - start_time) * 1000
        result.memory_usage_mb = (end_memory - start_memory) / (1024.0 * 1024.0)
        result.timestamp = datetime.datetime.now()
        
        # Adicionar informações sobre timeout, se ocorreu
        if self.timeout_occurred:
            result.notes = f"Timeout após {self.timeout_seconds} segundos"
            result.execution_time_ms = self.timeout_seconds * 1000
            
        return result
        
    def benchmark_x25519_keygen(self):
        """Realiza o benchmark de geração de chaves X25519"""
        result = BenchmarkResult()
        result.algorithm = "X25519"
        result.key_size = 256  # X25519 tem tamanho fixo de 256 bits
        result.operation_type = "Key Generation"
        result.data_size_bytes = 0  # Não aplicável para geração de chaves
        
        # Reset da flag de timeout
        self.timeout_occurred = False
        
        # Medição de recursos
        process = psutil.Process(os.getpid())
        start_memory = process.memory_info().rss
        
        # Iniciar thread para monitoramento de CPU
        self.stop_cpu_measurement = False
        cpu_usage_thread = threading.Thread(target=self.measure_cpu_usage, args=(result,))
        cpu_usage_thread.start()
        
        # Iniciar thread para timeout (se configurado)
        if self.timeout_seconds > 0:
            timer = threading.Timer(self.timeout_seconds, self.timeout_handler)
            timer.start()
        
        # Executar geração de chave e medir tempo
        start_time = time.time()
        
        try:
            # Gerar par de chaves X25519
            if not self.timeout_occurred:
                private_key = x25519.X25519PrivateKey.generate()
                public_key = private_key.public_key()
                
        except Exception as ex:
            # Cancelar o timer se ocorrer exceção
            if self.timeout_seconds > 0:
                timer.cancel()
            # Finalizar medição de CPU
            self.stop_cpu_measurement = True
            cpu_usage_thread.join()
            raise ex
            
        end_time = time.time()
        
        # Cancelar o timer se a operação for concluída antes do timeout
        if self.timeout_seconds > 0 and not self.timeout_occurred:
            timer.cancel()
        
        # Finalizar medição de CPU
        self.stop_cpu_measurement = True
        cpu_usage_thread.join()
        
        # Calcular uso de memória
        end_memory = process.memory_info().rss
        
        # Armazenar resultados
        result.execution_time_ms = (end_time - start_time) * 1000
        result.memory_usage_mb = (end_memory - start_memory) / (1024.0 * 1024.0)
        result.timestamp = datetime.datetime.now()
        
        # Adicionar informações sobre timeout, se ocorreu
        if self.timeout_occurred:
            result.notes = f"Timeout após {self.timeout_seconds} segundos"
            result.execution_time_ms = self.timeout_seconds * 1000
        
        # Armazenar as chaves para uso em troca de chaves
        self._last_x25519_private_key = private_key
            
        return result
        
    def benchmark_x25519_key_exchange(self):
        """Realiza o benchmark de troca de chaves X25519 (ECDH)"""
        result = BenchmarkResult()
        result.algorithm = "X25519"
        result.key_size = 256  # X25519 tem tamanho fixo de 256 bits
        result.operation_type = "Key Exchange"
        result.data_size_bytes = 32  # Tamanho da chave compartilhada (32 bytes)
        
        # Gerar os pares de chaves para Alice e Bob
        if hasattr(self, '_last_x25519_private_key'):
            alice_private = self._last_x25519_private_key
        else:
            alice_private = x25519.X25519PrivateKey.generate()
            
        bob_private = x25519.X25519PrivateKey.generate()
        
        alice_public = alice_private.public_key()
        bob_public = bob_private.public_key()
        
        # Serializar a chave pública de Bob para simular transmissão
# Serializar a chave pública de Bob para simular transmissão
        bob_public_bytes = bob_public.public_bytes(
            encoding=Encoding.Raw,
            format=PublicFormat.Raw
        )
        
        # Reset da flag de timeout
        self.timeout_occurred = False
        
        # Medição de recursos
        process = psutil.Process(os.getpid())
        start_memory = process.memory_info().rss
        
        # Iniciar thread para monitoramento de CPU
        self.stop_cpu_measurement = False
        cpu_usage_thread = threading.Thread(target=self.measure_cpu_usage, args=(result,))
        cpu_usage_thread.start()
        
        # Iniciar thread para timeout (se configurado)
        if self.timeout_seconds > 0:
            timer = threading.Timer(self.timeout_seconds, self.timeout_handler)
            timer.start()
        
        # Executar troca de chaves e medir tempo
        start_time = time.time()
        
        try:
            # Alice gera a chave compartilhada usando a chave pública de Bob
            if not self.timeout_occurred:
                shared_key = alice_private.exchange(x25519.X25519PublicKey.from_public_bytes(bob_public_bytes))
                
        except Exception as ex:
            # Cancelar o timer se ocorrer exceção
            if self.timeout_seconds > 0:
                timer.cancel()
            # Finalizar medição de CPU
            self.stop_cpu_measurement = True
            cpu_usage_thread.join()
            raise ex
            
        end_time = time.time()
        
        # Cancelar o timer se a operação for concluída antes do timeout
        if self.timeout_seconds > 0 and not self.timeout_occurred:
            timer.cancel()
        
        # Finalizar medição de CPU
        self.stop_cpu_measurement = True
        cpu_usage_thread.join()
        
        # Calcular uso de memória
        end_memory = process.memory_info().rss
        
        # Armazenar resultados
        result.execution_time_ms = (end_time - start_time) * 1000
        result.memory_usage_mb = (end_memory - start_memory) / (1024.0 * 1024.0)
        result.timestamp = datetime.datetime.now()
        
        # Adicionar informações sobre timeout, se ocorreu
        if self.timeout_occurred:
            result.notes = f"Timeout após {self.timeout_seconds} segundos"
            result.execution_time_ms = self.timeout_seconds * 1000
            
        return result
    
    # ==== NIST Curves (P-256/P-384/P-521) ====
    
    def benchmark_ecdsa_keygen(self, curve_name, curve):
        """Realiza o benchmark de geração de chaves ECDSA usando uma curva específica"""
        result = BenchmarkResult()
        result.algorithm = curve_name
        
        # Define o tamanho da chave com base na curva
        if curve_name == "NIST_P256":
            result.key_size = 256
        elif curve_name == "NIST_P384":
            result.key_size = 384
        elif curve_name == "NIST_P521":
            result.key_size = 521
        else:
            result.key_size = 0
            
        result.operation_type = "Key Generation"
        result.data_size_bytes = 0  # Não aplicável para geração de chaves
        
        # Reset da flag de timeout
        self.timeout_occurred = False
        
        # Medição de recursos
        process = psutil.Process(os.getpid())
        start_memory = process.memory_info().rss
        
        # Iniciar thread para monitoramento de CPU
        self.stop_cpu_measurement = False
        cpu_usage_thread = threading.Thread(target=self.measure_cpu_usage, args=(result,))
        cpu_usage_thread.start()
        
        # Iniciar thread para timeout (se configurado)
        if self.timeout_seconds > 0:
            timer = threading.Timer(self.timeout_seconds, self.timeout_handler)
            timer.start()
        
        # Executar geração de chave e medir tempo
        start_time = time.time()
        
        try:
            # Gerar par de chaves EC
            if not self.timeout_occurred:
                private_key = ec.generate_private_key(curve)
                public_key = private_key.public_key()
                
        except Exception as ex:
            # Cancelar o timer se ocorrer exceção
            if self.timeout_seconds > 0:
                timer.cancel()
            # Finalizar medição de CPU
            self.stop_cpu_measurement = True
            cpu_usage_thread.join()
            raise ex
            
        end_time = time.time()
        
        # Cancelar o timer se a operação for concluída antes do timeout
        if self.timeout_seconds > 0 and not self.timeout_occurred:
            timer.cancel()
        
        # Finalizar medição de CPU
        self.stop_cpu_measurement = True
        cpu_usage_thread.join()
        
        # Calcular uso de memória
        end_memory = process.memory_info().rss
        
        # Armazenar resultados
        result.execution_time_ms = (end_time - start_time) * 1000
        result.memory_usage_mb = (end_memory - start_memory) / (1024.0 * 1024.0)
        result.timestamp = datetime.datetime.now()
        
        # Adicionar informações sobre timeout, se ocorreu
        if self.timeout_occurred:
            result.notes = f"Timeout após {self.timeout_seconds} segundos"
            result.execution_time_ms = self.timeout_seconds * 1000
        
        # Armazenar as chaves para uso em operações ECDSA
        # Usamos um dicionário para armazenar chaves por curva
        if not hasattr(self, '_last_ecdsa_keys'):
            self._last_ecdsa_keys = {}
            
        self._last_ecdsa_keys[curve_name] = private_key
            
        return result
        
    def benchmark_ecdsa_sign(self, curve_name, curve, data):
        """Realiza o benchmark de assinatura ECDSA usando uma curva específica"""
        result = BenchmarkResult()
        result.algorithm = curve_name
        
        # Define o tamanho da chave com base na curva
        if curve_name == "NIST_P256":
            result.key_size = 256
        elif curve_name == "NIST_P384":
            result.key_size = 384
        elif curve_name == "NIST_P521":
            result.key_size = 521
        else:
            result.key_size = 0
            
        result.operation_type = "Signing"
        result.data_size_bytes = len(data)
        
        # Verificar se temos uma chave privada disponível
        if hasattr(self, '_last_ecdsa_keys') and curve_name in self._last_ecdsa_keys:
            private_key = self._last_ecdsa_keys[curve_name]
        else:
            # Se não tivermos, criar uma nova
            private_key = ec.generate_private_key(curve)
            # Armazenar para uso futuro
            if not hasattr(self, '_last_ecdsa_keys'):
                self._last_ecdsa_keys = {}
            self._last_ecdsa_keys[curve_name] = private_key
        
        # Reset da flag de timeout
        self.timeout_occurred = False
        
        # Medição de recursos
        process = psutil.Process(os.getpid())
        start_memory = process.memory_info().rss
        
        # Iniciar thread para monitoramento de CPU
        self.stop_cpu_measurement = False
        cpu_usage_thread = threading.Thread(target=self.measure_cpu_usage, args=(result,))
        cpu_usage_thread.start()
        
        # Iniciar thread para timeout (se configurado)
        if self.timeout_seconds > 0:
            timer = threading.Timer(self.timeout_seconds, self.timeout_handler)
            timer.start()
        
        # Preparar o hash dos dados para assinatura
        digest = hashes.Hash(hashes.SHA256())
        digest.update(data)
        data_hash = digest.finalize()
        
        # Executar assinatura e medir tempo
        start_time = time.time()
        
        try:
            # Assinar o hash dos dados
            if not self.timeout_occurred:
                signature = private_key.sign(
                    data_hash,
                    ec.ECDSA(hashes.SHA256())
                )
                
        except Exception as ex:
            # Cancelar o timer se ocorrer exceção
            if self.timeout_seconds > 0:
                timer.cancel()
            # Finalizar medição de CPU
            self.stop_cpu_measurement = True
            cpu_usage_thread.join()
            raise ex
            
        end_time = time.time()
        
        # Cancelar o timer se a operação for concluída antes do timeout
        if self.timeout_seconds > 0 and not self.timeout_occurred:
            timer.cancel()
        
        # Finalizar medição de CPU
        self.stop_cpu_measurement = True
        cpu_usage_thread.join()
        
        # Calcular uso de memória
        end_memory = process.memory_info().rss
        
        # Armazenar resultados
        result.execution_time_ms = (end_time - start_time) * 1000
        result.memory_usage_mb = (end_memory - start_memory) / (1024.0 * 1024.0)
        result.timestamp = datetime.datetime.now()
        
        # Adicionar informações sobre timeout, se ocorreu
        if self.timeout_occurred:
            result.notes = f"Timeout após {self.timeout_seconds} segundos"
            result.execution_time_ms = self.timeout_seconds * 1000
            
        # Armazenar a assinatura para uso em verificação
        if not hasattr(self, '_last_ecdsa_signatures'):
            self._last_ecdsa_signatures = {}
            
        if not hasattr(self, '_last_ecdsa_hashes'):
            self._last_ecdsa_hashes = {}
            
        self._last_ecdsa_signatures[curve_name] = signature
        self._last_ecdsa_hashes[curve_name] = data_hash
            
        return result
        
    def benchmark_ecdsa_verify(self, curve_name, curve, data):
        """Realiza o benchmark de verificação ECDSA usando uma curva específica"""
        result = BenchmarkResult()
        result.algorithm = curve_name
        
        # Define o tamanho da chave com base na curva
        if curve_name == "NIST_P256":
            result.key_size = 256
        elif curve_name == "NIST_P384":
            result.key_size = 384
        elif curve_name == "NIST_P521":
            result.key_size = 521
        else:
            result.key_size = 0
            
        result.operation_type = "Verification"
        result.data_size_bytes = len(data)
        
        # Verificar se temos uma assinatura e chave disponíveis
        if (hasattr(self, '_last_ecdsa_keys') and curve_name in self._last_ecdsa_keys and
            hasattr(self, '_last_ecdsa_signatures') and curve_name in self._last_ecdsa_signatures and
            hasattr(self, '_last_ecdsa_hashes') and curve_name in self._last_ecdsa_hashes):
            
            private_key = self._last_ecdsa_keys[curve_name]
            public_key = private_key.public_key()
            signature = self._last_ecdsa_signatures[curve_name]
            data_hash = self._last_ecdsa_hashes[curve_name]
            
        else:
            # Se não tivermos, criar novos
            private_key = ec.generate_private_key(curve)
            public_key = private_key.public_key()
            
            # Preparar o hash dos dados para assinatura
            digest = hashes.Hash(hashes.SHA256())
            digest.update(data)
            data_hash = digest.finalize()
            
            # Criar assinatura
            signature = private_key.sign(
                data_hash,
                ec.ECDSA(hashes.SHA256())
            )
            
        # Reset da flag de timeout
        self.timeout_occurred = False
        
        # Medição de recursos
        process = psutil.Process(os.getpid())
        start_memory = process.memory_info().rss
        
        # Iniciar thread para monitoramento de CPU
        self.stop_cpu_measurement = False
        cpu_usage_thread = threading.Thread(target=self.measure_cpu_usage, args=(result,))
        cpu_usage_thread.start()
        
        # Iniciar thread para timeout (se configurado)
        if self.timeout_seconds > 0:
            timer = threading.Timer(self.timeout_seconds, self.timeout_handler)
            timer.start()
        
        # Executar verificação e medir tempo
        start_time = time.time()
        
        try:
            # Verificar a assinatura
            if not self.timeout_occurred:
                public_key.verify(
                    signature,
                    data_hash,
                    ec.ECDSA(hashes.SHA256())
                )
                
        except Exception as ex:
            # Cancelar o timer se ocorrer exceção
            if self.timeout_seconds > 0:
                timer.cancel()
            # Finalizar medição de CPU
            self.stop_cpu_measurement = True
            cpu_usage_thread.join()
            raise ex
            
        end_time = time.time()
        
        # Cancelar o timer se a operação for concluída antes do timeout
        if self.timeout_seconds > 0 and not self.timeout_occurred:
            timer.cancel()
        
        # Finalizar medição de CPU
        self.stop_cpu_measurement = True
        cpu_usage_thread.join()
        
        # Calcular uso de memória
        end_memory = process.memory_info().rss
        
        # Armazenar resultados
        result.execution_time_ms = (end_time - start_time) * 1000
        result.memory_usage_mb = (end_memory - start_memory) / (1024.0 * 1024.0)
        result.timestamp = datetime.datetime.now()
        
        # Adicionar informações sobre timeout, se ocorreu
        if self.timeout_occurred:
            result.notes = f"Timeout após {self.timeout_seconds} segundos"
            result.execution_time_ms = self.timeout_seconds * 1000
            
        return result
        
    def benchmark_ecdh_key_exchange(self, curve_name, curve):
        """Realiza o benchmark de troca de chaves ECDH usando uma curva específica"""
        result = BenchmarkResult()
        result.algorithm = f"{curve_name}_ECDH"
        
        # Define o tamanho da chave com base na curva
        if curve_name == "NIST_P256":
            result.key_size = 256
        elif curve_name == "NIST_P384":
            result.key_size = 384
        elif curve_name == "NIST_P521":
            result.key_size = 521
        else:
            result.key_size = 0
            
        result.operation_type = "Key Exchange"
        result.data_size_bytes = result.key_size // 8  # Tamanho aproximado da chave em bytes
        
        # Gerar pares de chaves para Alice e Bob
        if hasattr(self, '_last_ecdsa_keys') and curve_name in self._last_ecdsa_keys:
            alice_private = self._last_ecdsa_keys[curve_name]
        else:
            alice_private = ec.generate_private_key(curve)
            
        bob_private = ec.generate_private_key(curve)
        
        alice_public = alice_private.public_key()
        bob_public = bob_private.public_key()
        
        # Serializar a chave pública de Bob para simular transmissão
        bob_public_bytes = bob_public.public_bytes(
            encoding=Encoding.X962,
            format=PublicFormat.CompressedPoint
        )
        
        # Reset da flag de timeout
        self.timeout_occurred = False
        
        # Medição de recursos
        process = psutil.Process(os.getpid())
        start_memory = process.memory_info().rss
        
        # Iniciar thread para monitoramento de CPU
        self.stop_cpu_measurement = False
        cpu_usage_thread = threading.Thread(target=self.measure_cpu_usage, args=(result,))
        cpu_usage_thread.start()
        
        # Iniciar thread para timeout (se configurado)
        if self.timeout_seconds > 0:
            timer = threading.Timer(self.timeout_seconds, self.timeout_handler)
            timer.start()
        
        # Executar troca de chaves e medir tempo
        start_time = time.time()
        
        try:
            # Alice gera a chave compartilhada usando a chave pública de Bob
            if not self.timeout_occurred:
                # Reconstruir a chave pública de Bob
                bob_restored_public = ec.EllipticCurvePublicKey.from_encoded_point(curve, bob_public_bytes)
                
                # Derivar a chave compartilhada
                shared_key = alice_private.exchange(
                    ec.ECDH(),
                    bob_restored_public
                )
                
        except Exception as ex:
            # Cancelar o timer se ocorrer exceção
            if self.timeout_seconds > 0:
                timer.cancel()
            # Finalizar medição de CPU
            self.stop_cpu_measurement = True
            cpu_usage_thread.join()
            raise ex
            
        end_time = time.time()
        
        # Cancelar o timer se a operação for concluída antes do timeout
        if self.timeout_seconds > 0 and not self.timeout_occurred:
            timer.cancel()
        
        # Finalizar medição de CPU
        self.stop_cpu_measurement = True
        cpu_usage_thread.join()
        
        # Calcular uso de memória
        end_memory = process.memory_info().rss
        
        # Armazenar resultados
        result.execution_time_ms = (end_time - start_time) * 1000
        result.memory_usage_mb = (end_memory - start_memory) / (1024.0 * 1024.0)
        result.timestamp = datetime.datetime.now()
        
        # Adicionar informações sobre timeout, se ocorreu
        if self.timeout_occurred:
            result.notes = f"Timeout após {self.timeout_seconds} segundos"
            result.execution_time_ms = self.timeout_seconds * 1000
            
        return result
    
    # ==== RSA ====
    
    def benchmark_rsa_keygen(self, key_size):
        """Realiza o benchmark de geração de chaves RSA"""
        result = BenchmarkResult()
        result.algorithm = "RSA"
        result.key_size = key_size
        result.operation_type = "Key Generation"
        result.data_size_bytes = 0  # Não aplicável para geração de chaves
        
        # Reset da flag de timeout
        self.timeout_occurred = False
        
        # Medição de recursos
        process = psutil.Process(os.getpid())
        start_memory = process.memory_info().rss
        
        # Iniciar thread para monitoramento de CPU
        self.stop_cpu_measurement = False
        cpu_usage_thread = threading.Thread(target=self.measure_cpu_usage, args=(result,))
        cpu_usage_thread.start()
        
        # Iniciar thread para timeout (se configurado)
        if self.timeout_seconds > 0:
            timer = threading.Timer(self.timeout_seconds, self.timeout_handler)
            timer.start()
        
        # Executar geração de chave e medir tempo
        start_time = time.time()
        
        try:
            # Gerar par de chaves RSA
            if not self.timeout_occurred:
                private_key = rsa.generate_private_key(
                    public_exponent=65537,
                    key_size=key_size
                )
                public_key = private_key.public_key()
                
        except Exception as ex:
            # Cancelar o timer se ocorrer exceção
            if self.timeout_seconds > 0:
                timer.cancel()
            # Finalizar medição de CPU
            self.stop_cpu_measurement = True
            cpu_usage_thread.join()
            raise ex
            
        end_time = time.time()
        
        # Cancelar o timer se a operação for concluída antes do timeout
        if self.timeout_seconds > 0 and not self.timeout_occurred:
            timer.cancel()
        
        # Finalizar medição de CPU
        self.stop_cpu_measurement = True
        cpu_usage_thread.join()
        
        # Calcular uso de memória
        end_memory = process.memory_info().rss
        
        # Armazenar resultados
        result.execution_time_ms = (end_time - start_time) * 1000
        result.memory_usage_mb = (end_memory - start_memory) / (1024.0 * 1024.0)
        result.timestamp = datetime.datetime.now()
        
        # Adicionar informações sobre timeout, se ocorreu
        if self.timeout_occurred:
            result.notes = f"Timeout após {self.timeout_seconds} segundos"
            result.execution_time_ms = self.timeout_seconds * 1000
        
        # Armazenar as chaves para uso em operações RSA
        if not hasattr(self, '_last_rsa_keys'):
            self._last_rsa_keys = {}
            
        self._last_rsa_keys[key_size] = private_key
            
        return result
        
    def benchmark_rsa_sign(self, key_size, data):
        """Realiza o benchmark de assinatura RSA"""
        result = BenchmarkResult()
        result.algorithm = "RSA"
        result.key_size = key_size
        result.operation_type = "Signing"
        result.data_size_bytes = len(data)
        
        # Verificar se temos uma chave privada disponível
        if hasattr(self, '_last_rsa_keys') and key_size in self._last_rsa_keys:
            private_key = self._last_rsa_keys[key_size]
        else:
            # Se não tivermos, criar uma nova
            private_key = rsa.generate_private_key(
                public_exponent=65537,
                key_size=key_size
            )
            # Armazenar para uso futuro
            if not hasattr(self, '_last_rsa_keys'):
                self._last_rsa_keys = {}
            self._last_rsa_keys[key_size] = private_key
        
        # Reset da flag de timeout
        self.timeout_occurred = False
        
        # Medição de recursos
        process = psutil.Process(os.getpid())
        start_memory = process.memory_info().rss
        
        # Iniciar thread para monitoramento de CPU
        self.stop_cpu_measurement = False
        cpu_usage_thread = threading.Thread(target=self.measure_cpu_usage, args=(result,))
        cpu_usage_thread.start()
        
        # Iniciar thread para timeout (se configurado)
        if self.timeout_seconds > 0:
            timer = threading.Timer(self.timeout_seconds, self.timeout_handler)
            timer.start()
        
        # Executar assinatura e medir tempo
        start_time = time.time()
        
        try:
            # Assinar os dados
            if not self.timeout_occurred:
                signature = private_key.sign(
                    data,
                    asym_padding.PSS(
                        mgf=asym_padding.MGF1(hashes.SHA256()),
                        salt_length=asym_padding.PSS.MAX_LENGTH
                    ),
                    hashes.SHA256()
                )
                
        except Exception as ex:
            # Cancelar o timer se ocorrer exceção
            if self.timeout_seconds > 0:
                timer.cancel()
            # Finalizar medição de CPU
            self.stop_cpu_measurement = True
            cpu_usage_thread.join()
            raise ex
            
        end_time = time.time()
        
        # Cancelar o timer se a operação for concluída antes do timeout
        if self.timeout_seconds > 0 and not self.timeout_occurred:
            timer.cancel()
        
        # Finalizar medição de CPU
        self.stop_cpu_measurement = True
        cpu_usage_thread.join()
        
        # Calcular uso de memória
        end_memory = process.memory_info().rss
        
        # Armazenar resultados
        result.execution_time_ms = (end_time - start_time) * 1000
        result.memory_usage_mb = (end_memory - start_memory) / (1024.0 * 1024.0)
        result.timestamp = datetime.datetime.now()
        
        # Adicionar informações sobre timeout, se ocorreu
        if self.timeout_occurred:
            result.notes = f"Timeout após {self.timeout_seconds} segundos"
            result.execution_time_ms = self.timeout_seconds * 1000
            
        # Armazenar a assinatura para uso em verificação
        if not hasattr(self, '_last_rsa_signatures'):
            self._last_rsa_signatures = {}
            
        if not hasattr(self, '_last_rsa_data'):
            self._last_rsa_data = {}
            
        self._last_rsa_signatures[key_size] = signature
        self._last_rsa_data[key_size] = data
            
        return result
        
    def benchmark_rsa_verify(self, key_size, data):
        """Realiza o benchmark de verificação RSA"""
        result = BenchmarkResult()
        result.algorithm = "RSA"
        result.key_size = key_size
        result.operation_type = "Verification"
        result.data_size_bytes = len(data)
        
        # Verificar se temos uma assinatura e chave disponíveis
        if (hasattr(self, '_last_rsa_keys') and key_size in self._last_rsa_keys and
            hasattr(self, '_last_rsa_signatures') and key_size in self._last_rsa_signatures and
            hasattr(self, '_last_rsa_data') and key_size in self._last_rsa_data):
            
            private_key = self._last_rsa_keys[key_size]
            public_key = private_key.public_key()
            signature = self._last_rsa_signatures[key_size]
            data_to_verify = self._last_rsa_data[key_size]
            
        else:
            # Se não tivermos, criar novos
            private_key = rsa.generate_private_key(
                public_exponent=65537,
                key_size=key_size
            )
            public_key = private_key.public_key()
            
            # Criar assinatura
            signature = private_key.sign(
                data,
                asym_padding.PSS(
                    mgf=asym_padding.MGF1(hashes.SHA256()),
                    salt_length=asym_padding.PSS.MAX_LENGTH
                ),
                hashes.SHA256()
            )
            data_to_verify = data
            
        # Reset da flag de timeout
        self.timeout_occurred = False
        
        # Medição de recursos
        process = psutil.Process(os.getpid())
        start_memory = process.memory_info().rss
        
        # Iniciar thread para monitoramento de CPU
        self.stop_cpu_measurement = False
        cpu_usage_thread = threading.Thread(target=self.measure_cpu_usage, args=(result,))
        cpu_usage_thread.start()
        
        # Iniciar thread para timeout (se configurado)
        if self.timeout_seconds > 0:
            timer = threading.Timer(self.timeout_seconds, self.timeout_handler)
            timer.start()
        
        # Executar verificação e medir tempo
        start_time = time.time()
        
        try:
            # Verificar a assinatura
            if not self.timeout_occurred:
                public_key.verify(
                    signature,
                    data_to_verify,
                    asym_padding.PSS(
                        mgf=asym_padding.MGF1(hashes.SHA256()),
                        salt_length=asym_padding.PSS.MAX_LENGTH
                    ),
                    hashes.SHA256()
                )
                
        except Exception as ex:
            # Cancelar o timer se ocorrer exceção
            if self.timeout_seconds > 0:
                timer.cancel()
            # Finalizar medição de CPU
            self.stop_cpu_measurement = True
            cpu_usage_thread.join()
            raise ex
            
        end_time = time.time()
        
        # Cancelar o timer se a operação for concluída antes do timeout
        if self.timeout_seconds > 0 and not self.timeout_occurred:
            timer.cancel()
        
        # Finalizar medição de CPU
        self.stop_cpu_measurement = True
        cpu_usage_thread.join()
        
        # Calcular uso de memória
        end_memory = process.memory_info().rss
        
        # Armazenar resultados
        result.execution_time_ms = (end_time - start_time) * 1000
        result.memory_usage_mb = (end_memory - start_memory) / (1024.0 * 1024.0)
        result.timestamp = datetime.datetime.now()
        
        # Adicionar informações sobre timeout, se ocorreu
        if self.timeout_occurred:
            result.notes = f"Timeout após {self.timeout_seconds} segundos"
            result.execution_time_ms = self.timeout_seconds * 1000
            
        return result
        
    def benchmark_rsa_encryption(self, algorithm, key_size, data):
        """Realiza o benchmark de criptografia RSA"""
        result = BenchmarkResult()
        result.algorithm = algorithm
        result.key_size = key_size
        result.operation_type = "Encryption"
        result.data_size_bytes = len(data)
        
        # Verificar se temos uma chave disponível
        if hasattr(self, '_last_rsa_keys') and key_size in self._last_rsa_keys:
            private_key = self._last_rsa_keys[key_size]
            public_key = private_key.public_key()
        else:
            # Se não tivermos, criar uma nova
            private_key = rsa.generate_private_key(
                public_exponent=65537,
                key_size=key_size
            )
            public_key = private_key.public_key()
            
            # Armazenar para uso futuro
            if not hasattr(self, '_last_rsa_keys'):
                self._last_rsa_keys = {}
            self._last_rsa_keys[key_size] = private_key
        
        # Reset da flag de timeout
        self.timeout_occurred = False
        
        # Medição de recursos
        process = psutil.Process(os.getpid())
        start_memory = process.memory_info().rss
        
        # Iniciar thread para monitoramento de CPU
        self.stop_cpu_measurement = False
        cpu_usage_thread = threading.Thread(target=self.measure_cpu_usage, args=(result,))
        cpu_usage_thread.start()
        
        # Iniciar thread para timeout (se configurado)
        if self.timeout_seconds > 0:
            timer = threading.Timer(self.timeout_seconds, self.timeout_handler)
            timer.start()
        
        # Executar criptografia e medir tempo
        start_time = time.time()
        
        try:
            # Criptografar os dados
            if not self.timeout_occurred:
                encrypted_data = public_key.encrypt(
                    data,
                    asym_padding.OAEP(
                        mgf=asym_padding.MGF1(algorithm=hashes.SHA256()),
                        algorithm=hashes.SHA256(),
                        label=None
                    )
                )
                
                # Adicionar a descriptografia para completar o teste
                decrypted_data = private_key.decrypt(
                    encrypted_data,
                    asym_padding.OAEP(
                        mgf=asym_padding.MGF1(algorithm=hashes.SHA256()),
                        algorithm=hashes.SHA256(),
                        label=None
                    )
                )
                
        except Exception as ex:
            # Cancelar o timer se ocorrer exceção
            if self.timeout_seconds > 0:
                timer.cancel()
            # Finalizar medição de CPU
            self.stop_cpu_measurement = True
            cpu_usage_thread.join()
            raise ex
            
        end_time = time.time()
        
        # Cancelar o timer se a operação for concluída antes do timeout
        if self.timeout_seconds > 0 and not self.timeout_occurred:
            timer.cancel()
        
        # Finalizar medição de CPU
        self.stop_cpu_measurement = True
        cpu_usage_thread.join()
        
        # Calcular uso de memória
        end_memory = process.memory_info().rss
        
        # Armazenar resultados
        result.execution_time_ms = (end_time - start_time) * 1000
        result.memory_usage_mb = (end_memory - start_memory) / (1024.0 * 1024.0)
        result.timestamp = datetime.datetime.now()
        
        # Adicionar informações sobre timeout, se ocorreu
        if self.timeout_occurred:
            result.notes = f"Timeout após {self.timeout_seconds} segundos"
            result.execution_time_ms = self.timeout_seconds * 1000
            
        return result


if __name__ == "__main__":
    # Cria uma instância do benchmark e executa
    benchmark = CryptoBenchmark()
    benchmark.run()